"""XLSX parser using openpyxl."""

from io import BytesIO

from openpyxl import load_workbook  # type: ignore[import-untyped]

from policymind.documents.models import DocumentBlock, ParsedDocument


class XLSXParser:
    """Extract worksheets and tables from .xlsx files."""

    supported_mime_types = frozenset(
        {
            "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        }
    )

    def parse(self, content: bytes, filename: str = "") -> ParsedDocument:
        wb = load_workbook(BytesIO(content), data_only=True)
        blocks: list[DocumentBlock] = []

        for sheet_name in wb.sheetnames:
            ws = wb[sheet_name]
            rows: list[str] = []
            row_count = 0
            for row in ws.iter_rows(values_only=True):
                cells = [str(c) if c is not None else "" for c in row]
                # Skip fully empty rows
                if any(cells):
                    rows.append(" | ".join(cells))
                    row_count += 1

            if rows:
                blocks.append(
                    DocumentBlock(
                        text=f"[Sheet: {sheet_name}]\n" + "\n".join(rows),
                        page_number=1,
                        block_type="table",
                    )
                )

        return ParsedDocument(
            title=filename or "Untitled",
            page_count=1,
            blocks=blocks,
            metadata={"parser": "openpyxl", "sheets": wb.sheetnames},
        )
